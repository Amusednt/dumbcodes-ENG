import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import ttk, messagebox

# Define o nome do arquivo da planilha
FILE_NAME = 'tarefas_aulas.xlsx'

# Função para criar a planilha inicial
def create_spreadsheet():
    wb = Workbook()
    ws = wb.active

    horarios = [
        "7h05 as 7h55", "7h55 as 8h45", "8h45 as 9h35", "9h35 as 10h",
        "10h as 10h50", "10h50 as 11h40", "11h40 as 12h30", 
        "13h25 as 14h15", "14h15 as 15h05", "15h05 as 15h55",
        "15h55 as 16h20", "16h20 as 17h10", "17h10 as 18h"
    ]

    dias_da_semana = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]

    for i, dia in enumerate(["Horários abaixo"] + dias_da_semana):
        ws.cell(row=1, column=i + 1, value=dia)

    for i, horario in enumerate(horarios):
        ws.cell(row=i + 2, column=1, value=horario)

    wb.save(FILE_NAME)

# Função para carregar a planilha
def load_spreadsheet():
    if os.path.exists(FILE_NAME):
        return load_workbook(FILE_NAME)
    else:
        create_spreadsheet()
        return load_workbook(FILE_NAME)

# Função para adicionar uma tarefa
def add_task(ws, dia, horario, descricao):
    dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    horarios = [
        "7h05 as 7h55", "7h55 as 8h45", "8h45 as 9h35", "9h35 as 10h",
        "10h as 10h50", "10h50 as 11h40", "11h40 as 12h30", 
        "13h25 as 14h15", "14h15 as 15h05", "15h05 as 15h55",
        "15h55 as 16h20", "16h20 as 17h10", "17h10 as 18h"
    ]
    
    dia_index = dias.index(dia) + 2  # Corrigido para começar em 2
    horario_index = horarios.index(horario) + 2
    
    # Verifica se já existe uma tarefa no horário e dia especificados
    if ws.cell(row=horario_index, column=dia_index).value:
        messagebox.showwarning("Atenção", "Já existe uma tarefa registrada neste horário.")
    else:
        ws.cell(row=horario_index, column=dia_index, value=descricao)

# Função para editar uma tarefa
def edit_task(ws, dia, horario, nova_descricao):
    dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    horarios = [
        "7h05 as 7h55", "7h55 as 8h45", "8h45 as 9h35", "9h35 as 10h",
        "10h as 10h50", "10h50 as 11h40", "11h40 as 12h30", 
        "13h25 as 14h15", "14h15 as 15h05", "15h05 as 15h55",
        "15h55 as 16h20", "16h20 as 17h10", "17h10 as 18h"
    ]
    
    dia_index = dias.index(dia) + 1
    horario_index = horarios.index(horario) + 2
    
    ws.cell(row=horario_index, column=dia_index, value=nova_descricao)

# Função para remover uma tarefa
def remove_task(ws, dia, horario):
    dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
    horarios = [
        "7h05 as 7h55", "7h55 as 8h45", "8h45 as 9h35", "9h35 as 10h",
        "10h as 10h50", "10h as 11h40", "11h40 as 12h30", 
        "13h25 as 14h15", "14h15 as 15h05", "15h05 as 15h55",
        "15h55 as 16h20", "16h20 as 17h10", "17h10 as 18h"
    ]
    
    dia_index = dias.index(dia) + 1
    horario_index = horarios.index(horario) + 2
    
    ws.cell(row=horario_index, column=dia_index, value='')

# Função para atualizar a planilha
def update_spreadsheet():
    wb = load_spreadsheet()
    ws = wb.active
    return wb, ws

# Função para adicionar tarefa via interface
def submit_task():
    dia = dia_combobox.get()
    horario = horario_combobox.get()
    descricao = descricao_entry.get()

    if dia and horario and descricao:
        wb, ws = update_spreadsheet()
        
        # Define os índices para dia e horário
        dias = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"]
        horarios = [
            "7h05 as 7h55", "7h55 as 8h45", "8h45 as 9h35", "9h35 as 10h",
            "10h as 10h50", "10h50 as 11h40", "11h40 as 12h30", 
            "13h25 as 14h15", "14h15 as 15h05", "15h05 as 15h55",
            "15h55 as 16h20", "16h20 as 17h10", "17h10 as 18h"
        ]
        
        dia_index = dias.index(dia) + 2  # Corrige para começar em 2
        horario_index = horarios.index(horario) + 2
        
        # Tenta adicionar a tarefa
        add_task(ws, dia, horario, descricao)
        
        # Salva a planilha apenas se a tarefa foi adicionada
        if ws.cell(row=horario_index, column=dia_index).value == descricao:
            wb.save(FILE_NAME)
            messagebox.showinfo("Sucesso", "Tarefa adicionada com sucesso!")
    else:
        messagebox.showwarning("Atenção", "Por favor, preencha todos os campos.")

# Função para criar a interface gráfica
def create_gui():
    global dia_combobox, horario_combobox, descricao_entry

    root = tk.Tk()
    root.title("Marcador de Tarefas")

    # Configurações da janela
    root.geometry("400x300")
    root.configure(bg="#f0f0f0")

    # Título
    title_label = tk.Label(root, text="Adicionar Tarefa", font=("Arial", 16), bg="#f0f0f0")
    title_label.pack(pady=10)

    # Combobox para dias da semana
    dia_label = tk.Label(root, text="Dia da Semana:", bg="#f0f0f0")
    dia_label.pack(pady=5)
    dia_combobox = ttk.Combobox(root, values=["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira"])
    dia_combobox.pack(pady=5)

    # Combobox para horários
    horario_label = tk.Label(root, text="Horário:", bg="#f0f0f0")
    horario_label.pack(pady=5)
    horario_combobox = ttk.Combobox(root, values=[
        "7h05 as 7h55", "7h55 as 8h45", "8h45 as 9h35", "9h35 as 10h",
        "10h as 10h50", "10h50 as 11h40", "11h40 as 12h30", 
        "13h25 as 14h15", "14h15 as 15h05", "15h05 as 15h55",
        "15h55 as 16h20", "16h20 as 17h10", "17h10 as 18h"
    ])
    horario_combobox.pack(pady=5)

    # Campo de entrada para descrição
    descricao_label = tk.Label(root, text="Descrição da Tarefa:", bg="#f0f0f0")
    descricao_label.pack(pady=5)
    descricao_entry = tk.Entry(root, width=30)
    descricao_entry.pack(pady=5)

    # Botão para adicionar tarefa
    submit_button = tk.Button(root, text="Adicionar Tarefa", command=submit_task, bg="#4CAF50", fg="white")
    submit_button.pack(pady=20)

    root.mainloop()

# Executa a interface gráfica
if __name__ == "__main__":
    create_gui()